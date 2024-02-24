from xlrd import open_workbook


def test_exel_file():
    reader = open_workbook("extract_dir/file_example_XLS_10.xls")

    sheet = reader.sheet_by_index(0)
# print(sheet.nrows)
# print(sheet.ncols)

# for nx in range(sheet.nrows):
#     print(sheet.row(nx))

    from openpyxl import load_workbook

    workbook = load_workbook("extract_dir/file_example_XLSX_50.xlsx")

    sheet = workbook.active

    print(sheet.cell(2, 2).value)

