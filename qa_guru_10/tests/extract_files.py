from zipfile import ZipFile
from xlrd import open_workbook
import os.path
import pypdf
def test_extract():

# exctrat zip

    extract_dir = 'extract_dir'

    with ZipFile("resources/Archive.zip") as reader:
        print(reader.namelist())
        reader.extractall(extract_dir)

# read xls

        reader = open_workbook("extract_dir/file_example_XLS_10.xls")

        sheet = reader.sheet_by_index(0)
        print(sheet.nrows)
        # print(sheet.ncols)

        # for nx in range(sheet.nrows):
        #     print(sheet.row(nx))

        from openpyxl import load_workbook

        workbook = load_workbook("extract_dir/file_example_XLSX_50.xlsx")

        sheet = workbook.active
        print(sheet.cell(2, 2).value)

# read pdf

        reader = pypdf.PdfReader("extract_dir/Python Testing with Pytest (Brian Okken) (1).pdf")
        print(reader.pages[1].extract_text())
        print(os.path.getsize("extract_dir/Python Testing with Pytest (Brian Okken) (1).pdf"))
